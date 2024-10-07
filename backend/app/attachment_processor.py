import base64
import fitz  # PyMuPDF
from openpyxl import load_workbook
import pytesseract
from PIL import Image
import io
from .exceptions import AttachmentProcessError
from .logger import setup_logger

logger = setup_logger(__name__)

class AttachmentProcessor:
    @staticmethod
    def process_attachment(attachment):
        try:
            content_type = attachment["content_type"]
            content = base64.b64decode(attachment["content"])

            if content_type == "application/pdf":
                return AttachmentProcessor._process_pdf(content)
            elif content_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]:
                return AttachmentProcessor._process_excel(content)
            elif content_type.startswith("image/"):
                return AttachmentProcessor._process_image(content)
            elif content_type == "text/plain":
                return AttachmentProcessor._process_text(content)
            else:
                logger.warning(f"Unsupported file type: {content_type}")
                return f"Unsupported file type: {content_type}"
        except Exception as e:
            logger.error(f"Error processing attachment: {str(e)}")
            raise AttachmentProcessError(f"Failed to process attachment: {str(e)}")

    @staticmethod
    def _process_pdf(content):
        try:
            doc = fitz.open(stream=content, filetype="pdf")
            text = ""
            for page in doc:
                page_text = page.get_text()
                if page_text.strip():
                    text += page_text
                else:  # If no text found, try OCR
                    pix = page.get_pixmap()
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    text += pytesseract.image_to_string(img, lang='kor+eng')  # Added Korean language support

            logger.info("Successfully processed PDF attachment")
            return AttachmentProcessor._clean_text(text)
        except Exception as e:
            logger.error(f"Error processing PDF: {str(e)}")
            return "PDF 처리 중 오류가 발생했습니다."

    @staticmethod
    def _process_excel(content):
        workbook = load_workbook(filename=io.BytesIO(content))
        text = ""
        for sheet in workbook.sheetnames:
            active_sheet = workbook[sheet]
            for row in active_sheet.iter_rows(values_only=True):
                text += " ".join(str(cell) for cell in row if cell is not None) + "\n"
        logger.info("Successfully processed Excel attachment")
        return AttachmentProcessor._clean_text(text)

    @staticmethod
    def _process_image(content):
        image = Image.open(io.BytesIO(content))
        text = pytesseract.image_to_string(image)
        logger.info("Successfully processed image attachment")
        return AttachmentProcessor._clean_text(text)

    @staticmethod
    def _process_text(content):
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            encodings = ['ascii', 'iso-8859-1', 'cp1252']
            for encoding in encodings:
                try:
                    text = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                logger.error("Failed to decode text file with all attempted encodings")
                return "Unable to decode text file"
        
        logger.info("Successfully processed text attachment")
        return AttachmentProcessor._clean_text(text)

    @staticmethod
    def _clean_text(text):
        cleaned_text = text.replace('\x00', '')
        cleaned_text = ''.join(char for char in cleaned_text if ord(char) >= 32 or char == '\n')
        return cleaned_text